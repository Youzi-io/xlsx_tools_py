"""测试复杂表头识别引擎。"""

import unittest
import os
import tempfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from src.core.header_detector import HeaderDetector, HeaderInfo, ColumnDef


def create_test_workbook(sheet_configs: list) -> openpyxl.Workbook:
    """
    辅助函数：创建测试用工作簿。

    sheet_configs: [(sheet_name, data_rows, header_styles, merged_ranges)]
      其中 data_rows 为二维列表（包含表头行和数据行）
    """
    wb = openpyxl.Workbook()
    for cfg in sheet_configs:
        name, data, styles, merges = cfg[0], cfg[1], cfg[2] if len(cfg) > 2 else None, cfg[3] if len(cfg) > 3 else None
        if wb.sheetnames == ['Sheet']:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)

        for row_idx, row_data in enumerate(data, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row_idx, col_idx, val)

        # 应用合并
        if merges:
            for merge_range in merges:
                ws.merge_cells(merge_range)

        # 应用样式（仅对表头）
        if styles:
            for style_info in styles:
                row, col, font_bold, fill_color = style_info
                cell = ws.cell(row, col)
                if font_bold:
                    cell.font = Font(bold=True)
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color,
                                            end_color=fill_color,
                                            fill_type="solid")

    return wb


class TestHeaderDetector(unittest.TestCase):
    """表头检测器测试。"""

    def test_single_row_header(self):
        """测试单行简单表头。"""
        data = [
            ["姓名", "年龄", "部门", "薪资"],
            ["张三", 28, "技术部", 15000],
            ["李四", 32, "市场部", 12000],
            ["王五", 25, "技术部", 18000],
        ]
        wb = create_test_workbook([("Sheet1", data, None, None)])
        ws = wb.active

        detector = HeaderDetector(ws)
        result = detector.detect()

        self.assertEqual(len(result.header_rows), 1)
        self.assertEqual(result.data_start_row, 1)
        self.assertEqual(len(result.columns), 4)
        self.assertIn("姓名", [c.flat_name for c in result.columns])
        self.assertIn("年龄", [c.flat_name for c in result.columns])

    def test_multi_row_header(self):
        """测试多行表头。"""
        data = [
            ["基本信息", "基本信息", "2024年度", "2024年度"],
            ["姓名", "部门", "Q1收入", "Q2收入"],
            ["张三", "技术部", 50000, 60000],
            ["李四", "市场部", 45000, 48000],
        ]
        wb = create_test_workbook([("Sheet1", data, None, None)])
        ws = wb.active

        detector = HeaderDetector(ws)
        result = detector.detect()

        self.assertEqual(len(result.header_rows), 2)
        self.assertEqual(result.data_start_row, 2)

        # 应该生成 4 个扁平化列名
        self.assertEqual(len(result.columns), 4)
        print("列名:", [c.flat_name for c in result.columns])
        print("层级:", [c.hierarchy for c in result.columns])

    def test_merged_cell_header(self):
        """测试合并单元格表头。"""
        data = [
            ["2024年度财务报表", "", "", ""],
            ["", "Q1", "Q2", "Q3"],
            ["收入", "支出", "收入", "支出"],
            ["张三", 100, 80, 120],
            ["李四", 90, 70, 110],
        ]
        merges = ["A1:D1"]  # 第1行合并 A-D
        wb = create_test_workbook([("Sheet1", data, None, merges)])
        ws = wb.active

        detector = HeaderDetector(ws)
        result = detector.detect()

        self.assertGreaterEqual(len(result.header_rows), 2)
        print("合并表头 - 列名:", [c.flat_name for c in result.columns])
        print("合并表头 - 层级:", [c.hierarchy for c in result.columns])

    def test_styled_header_detection(self):
        """测试基于格式的表头检测。"""
        data = [
            ["Name", "Age", "Salary", "Department"],
            ["Alice", 30, 80000, "Engineering"],
            ["Bob", 35, 90000, "Marketing"],
        ]
        styles = [
            (1, 1, True, "4472C4"),  # 加粗 + 蓝色背景
            (1, 2, True, "4472C4"),
            (1, 3, True, "4472C4"),
            (1, 4, True, "4472C4"),
        ]
        wb = create_test_workbook([("Sheet1", data, styles, None)])
        ws = wb.active

        detector = HeaderDetector(ws)
        result = detector.detect()

        self.assertEqual(result.data_start_row, 1)
        self.assertGreater(result.confidence, 0.5)

    def test_empty_sheet(self):
        """测试空 Sheet。"""
        wb = create_test_workbook([("Empty", [], None, None)])
        ws = wb.active

        detector = HeaderDetector(ws)
        result = detector.detect()

        # 空 Sheet：无内容，data_start_row 应为 0（无数据行）
        self.assertEqual(result.data_start_row, 0)
        self.assertEqual(len(result.columns), 0)

    def test_column_hierarchy_paths(self):
        """测试层级路径正确性。"""
        data = [
            ["类别A", "类别A", "类别B"],
            ["子项1", "子项2", "子项3"],
            [10, 20, 30],
        ]
        wb = create_test_workbook([("Sheet1", data, None, None)])
        ws = wb.active

        detector = HeaderDetector(ws)
        result = detector.detect()

        for col in result.columns:
            self.assertGreater(len(col.hierarchy), 0)
            print(f"  {col.original_excel_col}: {col.hierarchy} -> {col.flat_name}")


if __name__ == "__main__":
    unittest.main()
