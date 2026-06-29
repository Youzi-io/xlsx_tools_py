"""合并单元格解析工具 — 将 openpyxl merged_cells 展开为完整的行列映射。"""

from typing import Dict, List, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


def parse_merged_cells(ws: Worksheet) -> Dict[Tuple[int, int], Cell]:
    """
    解析工作表中的合并单元格，将合并区域中每个单元格都映射到左上角的值。

    返回: {(row, col): source_cell} 的字典。
    对于未合并的单元格，映射到自身。
    行和列使用 1-based 索引（与 openpyxl 一致）。
    """
    # 先获取所有单元格的值
    cell_map: Dict[Tuple[int, int], Cell] = {}

    # 遍历所有有值的单元格
    for row in ws.iter_rows():
        for cell in row:
            cell_map[(cell.row, cell.col_idx)] = cell

    # 处理合并单元格区域
    merged_map: Dict[Tuple[int, int], Cell] = {}
    for merged_range in ws.merged_cells.ranges:
        # 获取左上角的值
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        # 创建一个虚拟 Cell 来保存值（避免创建实际 Cell 对象的复杂性）
        top_cell = cell_map.get((merged_range.min_row, merged_range.min_col))

        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if top_cell is not None:
                    merged_map[(row, col)] = top_cell

    # 合并：普通单元格 + 合并区域的映射
    full_map = {**cell_map, **merged_map}
    return full_map


def get_merged_region_map(ws: Worksheet) -> Dict[Tuple[int, int], str]:
    """
    获取每个单元格所属的合并区域标识。

    返回: {(row, col): region_key}，region_key 格式如 "A1:B2"
    未合并的单元格映射到自身坐标。
    """
    region_map: Dict[Tuple[int, int], str] = {}

    for merged_range in ws.merged_cells.ranges:
        region_key = str(merged_range)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                region_map[(row, col)] = region_key

    return region_map


def expand_merged_value(ws: Worksheet, row: int, col: int) -> object:
    """
    获取指定坐标的实际值，自动处理合并单元格。
    如果该坐标属于一个合并区域，返回左上角的值。
    """
    # 检查该坐标是否在合并区域中
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    # 不在任何合并区域中，直接返回
    return ws.cell(row, col).value


def is_in_merged_range(ws: Worksheet, row: int, col: int) -> bool:
    """检查指定坐标是否属于某个合并单元格区域。"""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False
