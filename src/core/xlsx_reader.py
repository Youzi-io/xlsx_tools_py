"""XLSX 读取器 — 封装 openpyxl，提供安全的文件读取和 Sheet 信息获取。"""

import os
from dataclasses import dataclass
from typing import Dict, List, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


@dataclass
class SheetInfo:
    """Sheet 的基本信息。"""
    name: str
    index: int
    rows: int
    cols: int
    has_merged_cells: bool
    merged_cells_count: int


@dataclass
class FileInfo:
    """XLSX 文件的基本信息。"""
    filepath: str
    filename: str
    sheet_count: int
    sheets: List[SheetInfo]


class XlsxReader:
    """
    XLSX 文件读取器。

    功能:
    - 安全打开 xlsx 文件（只读模式）
    - 获取所有 Sheet 的名称和基本信息
    - 按行读取数据
    - 获取合并单元格信息
    - 检测前 N 行的格式特征
    """

    def __init__(self, filepath: str):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"文件不存在: {filepath}")
        if not filepath.lower().endswith(".xlsx"):
            raise ValueError(f"仅支持 .xlsx 文件，当前文件: {filepath}")

        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        self._workbook: Optional[openpyxl.Workbook] = None
        self._active_sheet: Optional[Worksheet] = None
        self._active_sheet_name: Optional[str] = None

    # ── 上下文管理 ──────────────────────────────

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def open(self):
        """打开工作簿（data_only 模式读取计算值）。"""
        if self._workbook is None:
            self._workbook = openpyxl.load_workbook(
                self.filepath, data_only=True
            )

    def close(self):
        """关闭工作簿。"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._active_sheet = None
            self._active_sheet_name = None

    # ── Sheet 操作 ──────────────────────────────

    def get_sheet_names(self) -> List[str]:
        """获取所有 Sheet 名称。"""
        self._ensure_open()
        return self._workbook.sheetnames

    def get_file_info(self) -> FileInfo:
        """获取文件的完整信息。"""
        self._ensure_open()
        sheets = []
        for idx, name in enumerate(self._workbook.sheetnames):
            ws = self._workbook[name]
            merged_count = len(ws.merged_cells.ranges)
            sheets.append(SheetInfo(
                name=name,
                index=idx,
                rows=ws.max_row or 0,
                cols=ws.max_column or 0,
                has_merged_cells=merged_count > 0,
                merged_cells_count=merged_count,
            ))
        return FileInfo(
            filepath=self.filepath,
            filename=self.filename,
            sheet_count=len(sheets),
            sheets=sheets,
        )

    def select_sheet(self, name: str) -> Worksheet:
        """选择当前操作的 Sheet。"""
        self._ensure_open()
        if name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' 不存在。可用: {self._workbook.sheetnames}")
        self._active_sheet = self._workbook[name]
        self._active_sheet_name = name
        return self._active_sheet

    def get_active_sheet(self) -> Worksheet:
        """获取当前 Sheet（若未选择则使用第一个）。"""
        self._ensure_open()
        if self._active_sheet is None:
            self.select_sheet(self._workbook.sheetnames[0])
        return self._active_sheet

    # ── 数据读取 ────────────────────────────────

    def read_rows(self, start_row: int = 1, end_row: Optional[int] = None,
                  min_col: int = 1, max_col: Optional[int] = None) -> List[List[object]]:
        """
        按行读取数据，返回二维列表。

        参数:
            start_row: 起始行 (1-based)
            end_row: 结束行，None 表示到最后
        """
        ws = self.get_active_sheet()
        if end_row is None:
            end_row = ws.max_row
        if max_col is None:
            max_col = ws.max_column

        data = []
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(min_col, max_col + 1):
                row_data.append(ws.cell(row_idx, col_idx).value)
            data.append(row_data)
        return data

    def read_header_rows(self, n_rows: int = 10) -> List[List[object]]:
        """读取前 N 行（用于表头分析）。"""
        ws = self.get_active_sheet()
        end = min(n_rows, ws.max_row)
        return self.read_rows(1, end)

    def read_data_rows(self, start_row: int, end_row: Optional[int] = None) -> List[List[object]]:
        """读取数据行（从指定行开始）。"""
        return self.read_rows(start_row, end_row)

    # ── 格式特征提取 ────────────────────────────

    def get_row_styles(self, row_idx: int) -> List[Dict]:
        """
        获取指定行每个单元格的样式信息。

        返回每个单元格的样式字典: {bold, font_size, fill_color, alignment, border}
        """
        ws = self.get_active_sheet()
        styles = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            font = cell.font
            fill = cell.fill
            styles.append({
                "bold": font.bold if font else False,
                "font_size": font.size if font else None,
                "fill_fgColor": fill.fgColor.rgb if fill and fill.fgColor else None,
                "fill_bgColor": fill.bgColor.rgb if fill and fill.bgColor else None,
                "has_fill": fill.patternType is not None if fill else False,
                "col_letter": get_column_letter(col_idx),
            })
        return styles

    def get_merged_ranges(self) -> List[str]:
        """获取当前 Sheet 的所有合并区域（字符串格式）。"""
        ws = self.get_active_sheet()
        return [str(r) for r in ws.merged_cells.ranges]

    def get_merged_ranges_raw(self) -> list:
        """获取当前 Sheet 的所有合并区域（openpyxl 对象）。"""
        ws = self.get_active_sheet()
        return list(ws.merged_cells.ranges)

    # ── 内部方法 ────────────────────────────────

    def _ensure_open(self):
        if self._workbook is None:
            raise RuntimeError("工作簿未打开。请先调用 open() 或使用 with 语句。")

    @property
    def workbook(self) -> openpyxl.Workbook:
        self._ensure_open()
        return self._workbook
