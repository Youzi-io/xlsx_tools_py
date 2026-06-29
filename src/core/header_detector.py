"""
复杂表头识别引擎。

核心功能: 自动识别 xlsx 中多行表头、合并单元格表头，生成扁平化的列名。

识别策略（多维度评分）:
1. 格式特征 — 表头行通常有加粗/背景色/边框
2. 数据类型 — 表头通常是字符串，数据行通常是数值/日期/混合类型
3. 空值比例 — 表头行可能因合并单元格而有较多空值
4. 合并单元格 — 合并区域的值需要展开填充
5. 内容模式 — 表头行更可能包含"合计"、"总计"、"序号"等关键词
"""

from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from src.utils.merger import expand_merged_value
from src.utils.formatter import sanitize_column_name, flatten_header, infer_column_type


@dataclass
class ColumnDef:
    """单个列的定义。"""
    col_index: int                  # 列索引 (0-based)
    flat_name: str                  # 扁平化列名，如 "2024年度_Q1_收入"
    hierarchy: List[str] = field(default_factory=list)  # 层级路径
    original_excel_col: str = ""    # Excel 列字母 "A"
    is_merged: bool = False         # 是否来自合并单元格
    dtype: str = "string"           # 推断的数据类型


@dataclass
class HeaderInfo:
    """表头识别结果。"""
    header_rows: List[int]          # 表头所占行号 (0-based 相对于数据起始行)
    data_start_row: int             # 数据起始行号 (0-based)
    columns: List[ColumnDef]         # 列定义列表
    total_rows: int                 # Sheet 总行数
    total_cols: int                 # Sheet 总列数
    sheet_name: str = ""            # Sheet 名称
    confidence: float = 0.0         # 识别置信度 (0-1)


class HeaderDetector:
    """
    复杂表头检测器。

    用法:
        detector = HeaderDetector(ws, max_scan_rows=20)
        header_info = detector.detect()
    """

    # 表头关键词（出现这些词的行更有可能是表头）
    HEADER_KEYWORDS = {
        "序号", "编号", "名称", "日期", "时间", "金额", "数量", "备注", "说明",
        "部门", "姓名", "性别", "年龄", "地址", "电话", "合计", "总计", "小计",
        "单位", "单价", "总价", "收入", "支出", "利润", "占比", "同比", "环比",
        "id", "name", "date", "amount", "total", "qty", "price", "desc",
        "NO", "序号", "项目", "类别", "类型", "状态",
    }

    def __init__(self, ws: Worksheet, max_scan_rows: int = 50):
        self.ws = ws
        self.max_scan_rows = min(max_scan_rows, ws.max_row)
        self.max_col = ws.max_column
        self.merged_regions = list(ws.merged_cells.ranges)

    def detect(self) -> HeaderInfo:
        """
        执行表头检测，返回 HeaderInfo。

        算法:
        1. 扫描前 N 行，为每行计算"表头得分"
        2. 找到得分最高的连续行作为表头区域
        3. 构建层级列名
        4. 返回结果
        """
        # 快速检查是否为空 Sheet（前 5 行全空则视为空表）
        has_any_data = any(
            self.ws.cell(r, c).value is not None
            for r in range(1, min(self.ws.max_row, 5) + 1)
            for c in range(1, self.max_col + 1)
        )
        if not has_any_data:
            return HeaderInfo(
                header_rows=[],
                data_start_row=0,
                columns=[],
                total_rows=self.ws.max_row,
                total_cols=self.max_col,
                sheet_name=self.ws.title,
                confidence=1.0,
            )

        scan_rows = min(self.max_scan_rows, self.ws.max_row)

        # 步骤 1: 逐行评分
        row_scores = []
        for row_idx in range(1, scan_rows + 1):
            score = self._score_row_as_header(row_idx)
            row_scores.append((row_idx, score))

        # 步骤 2: 找到表头边界
        header_end_row = self._find_header_boundary(row_scores)
        # 表头至少占 1 行
        if header_end_row < 1:
            header_end_row = 1

        # 步骤 3: 确认表头起始行（从第 1 行到边界行中，得分超阈值的连续行）
        header_rows = list(range(1, header_end_row + 1))

        # 步骤 4: 构建列定义
        columns = self._build_column_defs(header_rows)

        # 步骤 5: 计算置信度
        confidence = self._calc_confidence(row_scores, header_rows)

        return HeaderInfo(
            header_rows=[r - 1 for r in header_rows],  # 转为 0-based
            data_start_row=header_end_row,              # 数据起始行 (0-based，即数据的第 1 行)
            columns=columns,
            total_rows=self.ws.max_row,
            total_cols=self.ws.max_column,
            sheet_name=self.ws.title,
            confidence=confidence,
        )

    # ── 行评分 ──────────────────────────────────

    def _score_row_as_header(self, row_idx: int) -> float:
        """
        综合评分某行是否为表头行。返回 0-1 之间的得分。

        考虑维度:
        - 格式得分 (0-0.4): 加粗、背景色、字号
        - 类型得分 (0-0.3): 字符串比例高
        - 关键词得分 (0-0.2): 包含表头关键词
        - 合并单元格得分 (0-0.1): 该行有合并单元格
        """
        format_score = self._calc_format_score(row_idx)
        type_score = self._calc_type_score(row_idx)
        keyword_score = self._calc_keyword_score(row_idx)
        merge_score = self._calc_merge_score(row_idx)

        total = (format_score * 0.35 + type_score * 0.35 +
                 keyword_score * 0.20 + merge_score * 0.10)
        return min(total, 1.0)

    def _calc_format_score(self, row_idx: int) -> float:
        """计算样式得分：加粗、背景色。"""
        try:
            score = 0.0
            non_empty = 0
            for col_idx in range(1, self.max_col + 1):
                cell = self.ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    non_empty += 1
                    if cell.font and cell.font.bold:
                        score += 1.0
                    if cell.fill and cell.fill.patternType is not None:
                        score += 0.5
            if non_empty == 0:
                return 0.0
            return min(score / non_empty, 1.0)
        except Exception:
            return 0.0

    def _calc_type_score(self, row_idx: int) -> float:
        """计算类型得分：字符串比例越高，越像表头。"""
        values = []
        for col_idx in range(1, self.max_col + 1):
            v = expand_merged_value(self.ws, row_idx, col_idx)
            values.append(v)

        non_null = [v for v in values if v is not None]
        if not non_null:
            return 0.0

        str_count = sum(1 for v in non_null if isinstance(v, str) and len(str(v).strip()) > 0)
        return str_count / len(non_null) if non_null else 0.0

    def _calc_keyword_score(self, row_idx: int) -> float:
        """计算关键词得分。"""
        matches = 0
        non_empty = 0
        for col_idx in range(1, self.max_col + 1):
            v = self.ws.cell(row_idx, col_idx).value
            if v is not None:
                non_empty += 1
                s = str(v).strip().lower()
                for kw in self.HEADER_KEYWORDS:
                    if kw.lower() in s:
                        matches += 1
                        break
        if non_empty == 0:
            return 0.0
        return min(matches / non_empty, 1.0)

    def _calc_merge_score(self, row_idx: int) -> float:
        """计算合并单元格得分：该行有多少单元格处于合并区域中。"""
        merged_count = 0
        for col_idx in range(1, self.max_col + 1):
            for mr in self.merged_regions:
                if (mr.min_row <= row_idx <= mr.max_row and
                        mr.min_col <= col_idx <= mr.max_col):
                    merged_count += 1
                    break
        return min(merged_count / max(self.max_col, 1), 1.0)

    # ── 边界判定 ────────────────────────────────

    def _find_header_boundary(self, row_scores: List[Tuple[int, float]]) -> int:
        """
        根据行得分找到表头边界。

        策略:
        - 找到连续高分区域
        - 边界条件是：当前行得分低于阈值 0.3，或者类型得分突变（数值比例突然升高）
        """
        # 使用滑动窗口找最高平均得分的连续区域
        best_start = 1
        best_end = 1
        best_avg = 0.0

        # 对前 20 行中所有可能的连续区域评分
        max_window = min(len(row_scores), 10)  # 表头不可能超过 10 行
        for start in range(len(row_scores)):
            for end in range(start, min(start + max_window, len(row_scores))):
                window_scores = [s for _, s in row_scores[start:end + 1]]
                avg = sum(window_scores) / len(window_scores)

                # 额外加分：连续行得分的稳定性
                if len(window_scores) > 1:
                    # 得分波动小更好
                    variance = sum((s - avg) ** 2 for s in window_scores) / len(window_scores)
                    stability_bonus = max(0, 0.1 - variance * 2)
                    avg += stability_bonus

                if avg > best_avg:
                    best_avg = avg
                    best_start = row_scores[start][0]
                    best_end = row_scores[end][0]

        # 如果最佳结束行得分太低，截断
        threshold = 0.25
        result_end = best_end
        for i in range(best_start - 1, min(best_end, len(row_scores))):
            idx, score = row_scores[i]
            if idx >= best_start and score < threshold:
                result_end = idx - 1
                break

        return result_end

    # ── 列定义构建 ──────────────────────────────

    def _build_column_defs(self, header_rows: List[int]) -> List[ColumnDef]:
        """
        从多行表头构建层级列定义。

        对每一列，从所有表头行中提取值，构建层级路径，
        然后扁平化生成唯一列名。
        """
        from openpyxl.utils import get_column_letter

        columns = []
        used_flat_names: Dict[str, int] = {}  # 用于处理重名

        for col_idx in range(1, self.max_col + 1):
            hierarchy = []
            is_merged = False

            for row_idx in header_rows:
                val = expand_merged_value(self.ws, row_idx, col_idx)
                hierarchy.append(sanitize_column_name(val))

                # 检查是否来自合并单元格
                for mr in self.merged_regions:
                    if (mr.min_row <= row_idx <= mr.max_row and
                            mr.min_col <= col_idx <= mr.max_col and
                            (mr.max_row > mr.min_row or mr.max_col > mr.min_col)):
                        is_merged = True
                        break

            # 去重和清理层级
            cleaned = self._clean_hierarchy(hierarchy)

            # 生成扁平化列名
            flat_name = flatten_header(cleaned)

            # 处理重名
            if flat_name in used_flat_names:
                used_flat_names[flat_name] += 1
                flat_name = f"{flat_name}_{used_flat_names[flat_name]}"
            else:
                used_flat_names[flat_name] = 0

            col_def = ColumnDef(
                col_index=col_idx - 1,  # 0-based
                flat_name=flat_name,
                hierarchy=cleaned,
                original_excel_col=get_column_letter(col_idx),
                is_merged=is_merged,
            )
            columns.append(col_def)

        return columns

    def _clean_hierarchy(self, parts: List[str]) -> List[str]:
        """清理层级路径：去除连续重复和空值。"""
        cleaned = []
        for p in parts:
            if not p or p == "Unnamed":
                continue
            if cleaned and p == cleaned[-1]:
                # 连续重复的去重（合并单元格导致）
                continue
            cleaned.append(p)

        # 如果全部为空，至少给一个默认名
        if not cleaned:
            cleaned = ["Column"]

        return cleaned

    # ── 置信度计算 ──────────────────────────────

    def _calc_confidence(self, row_scores, header_rows) -> float:
        """计算表头识别的置信度。"""
        if not header_rows or not row_scores:
            return 0.0

        # 表头行的平均得分
        header_scores = [s for r, s in row_scores if r in header_rows]
        if not header_scores:
            return 0.0

        avg_header_score = sum(header_scores) / len(header_scores)

        # 第一行数据（表头下一行）得分应较低
        data_row = header_rows[-1] + 1
        if data_row <= len(row_scores):
            data_score = row_scores[data_row - 1][1]  # row_scores 是 0-based
            # 差异越大置信度越高
            diff = avg_header_score - data_score
            confidence = 0.5 + diff * 0.5
        else:
            confidence = avg_header_score * 0.8

        return max(0.0, min(confidence, 1.0))
